import csv
import codecs
import os
import zipfile
from collections.abc import Iterable, Iterator, Sequence
from decimal import Decimal, InvalidOperation
from itertools import chain, islice
from pathlib import Path

import openpyxl

from apps.datasources.base import BaseDataSourceAdapter
from apps.datasources.limits import datasource_limit


class CSVValidationError(ValueError):
    pass


class CSVAdapter(BaseDataSourceAdapter):
    REQUIRED_COLUMNS = ['article', 'name', 'price', 'stock_qty']
    OPTIONAL_COLUMNS = ['brand', 'category', 'condition', 'oem_numbers', 'cross_numbers', 'description']
    MAX_HEADER_SCAN_ROWS = 50
    MAX_HEADER_DEPTH = 2
    HEADER_DATA_SAMPLE_ROWS = 5
    CSV_SNIFF_BYTES = 64 * 1024
    READ_CHUNK_BYTES = 64 * 1024

    def fetch_changes(self, since=None, limit=500, offset=0) -> list[dict]:
        raise NotImplementedError('CSVAdapter использует process_uploaded_file()')

    def test_connection(self) -> bool:
        return True

    def get_display_name(self) -> str:
        return 'CSV/Excel'

    def process_uploaded_file(self, file_path: str) -> list[dict]:
        return self._normalize(self._read_rows(file_path))

    def preview(self, file_path: str, rows: int = 10) -> dict:
        all_rows = self._read_rows(file_path)
        headers = list(all_rows[0].keys()) if all_rows else []
        return {
            'headers': headers,
            'rows': all_rows[:rows],
            'total_rows': len(all_rows),
        }

    def _read_rows(self, file_path: str) -> list[dict]:
        """Выбирает парсер по расширению; неподдерживаемый формат → понятная ошибка (а не 500)."""
        self._validate_file_size(file_path)
        suffix = Path(file_path).suffix.lower()
        if suffix == '.xlsx':
            return self._read_xlsx(file_path)
        if suffix == '.xls':
            return self._read_xls(file_path)
        if suffix in ('.csv', '.txt', ''):
            return self._read_csv(file_path)
        raise CSVValidationError(
            f'Формат "{suffix}" не поддерживается. Загрузите файл .csv, .xls или .xlsx.'
        )

    @staticmethod
    def _validate_file_size(file_path: str) -> None:
        max_bytes = datasource_limit('DATASOURCE_UPLOAD_MAX_BYTES')
        try:
            file_size = os.path.getsize(file_path)
        except OSError as exc:
            raise CSVValidationError('Не удалось прочитать загруженный файл.') from exc
        if file_size > max_bytes:
            raise CSVValidationError(
                'Размер файла превышает допустимый лимит '
                f'{max_bytes} байт.'
            )

    def _read_csv(self, file_path: str) -> list[dict]:
        encoding = self._detect_csv_encoding(file_path)
        try:
            with open(file_path, encoding=encoding, newline='') as source:
                sample = source.read(self.CSV_SNIFF_BYTES)
                source.seek(0)
                reader = csv.reader(
                    source,
                    delimiter=self._detect_delimiter(sample),
                    strict=True,
                )
                return self._rows_to_dicts(reader)
        except (csv.Error, UnicodeError) as exc:
            raise CSVValidationError('CSV повреждён или имеет некорректный формат.') from exc

    def _detect_csv_encoding(self, file_path: str) -> str:
        decoder = codecs.getincrementaldecoder('utf-8-sig')()
        try:
            with open(file_path, 'rb') as source:
                while chunk := source.read(self.READ_CHUNK_BYTES):
                    decoder.decode(chunk, final=False)
                decoder.decode(b'', final=True)
        except UnicodeDecodeError:
            return 'cp1251'
        return 'utf-8-sig'

    @staticmethod
    def _detect_delimiter(text: str) -> str:
        """Определяет разделитель CSV (',', ';' или таб) — русский Excel часто пишет ';'."""
        lines = text.splitlines()
        try:
            return csv.Sniffer().sniff('\n'.join(lines[:10]), delimiters=',;\t').delimiter
        except csv.Error:
            first = lines[0] if lines else ''
            counts = {d: first.count(d) for d in (',', ';', '\t')}
            return max(counts, key=counts.get) if any(counts.values()) else ','

    def _read_xlsx(self, file_path: str) -> list[dict]:
        self._validate_xlsx_archive(file_path)
        try:
            wb = openpyxl.load_workbook(
                file_path,
                read_only=True,
                data_only=True,
                keep_links=False,
            )
        except Exception as exc:
            raise CSVValidationError(
                'Файл повреждён или не является настоящим .xlsx. '
                'Пересохраните его в Excel как «Книга Excel (.xlsx)».'
            ) from exc
        try:
            ws = wb.active
            self._validate_sheet_dimensions(ws.max_row, ws.max_column)
            return self._rows_to_dicts(ws.iter_rows(values_only=True))
        finally:
            wb.close()

    @staticmethod
    def _validate_xlsx_archive(file_path: str) -> None:
        max_expanded_bytes = datasource_limit(
            'DATASOURCE_XLSX_MAX_UNCOMPRESSED_BYTES',
        )
        max_entries = datasource_limit('DATASOURCE_XLSX_MAX_ARCHIVE_ENTRIES')
        try:
            with zipfile.ZipFile(file_path) as archive:
                expanded_bytes = 0
                members = archive.infolist()
                if len(members) > max_entries:
                    raise CSVValidationError(
                        'Количество файлов внутри .xlsx превышает допустимый лимит '
                        f'{max_entries}.',
                    )
                for member in members:
                    if member.flag_bits & 0x1:
                        raise CSVValidationError('Зашифрованные .xlsx файлы не поддерживаются.')
                    expanded_bytes += member.file_size
                    if expanded_bytes > max_expanded_bytes:
                        raise CSVValidationError(
                            'Распакованный размер .xlsx превышает допустимый лимит '
                            f'{max_expanded_bytes} байт.'
                        )
        except CSVValidationError:
            raise
        except (OSError, zipfile.BadZipFile) as exc:
            raise CSVValidationError(
                'Файл повреждён или не является настоящим .xlsx. '
                'Пересохраните его в Excel как «Книга Excel (.xlsx)».'
            ) from exc

    def _read_xls(self, file_path: str) -> list[dict]:
        try:
            import xlrd
        except ImportError as exc:  # pragma: no cover — xlrd объявлен в requirements
            raise CSVValidationError('Чтение .xls временно недоступно. Сохраните файл как .xlsx.') from exc
        try:
            wb = xlrd.open_workbook(file_path)
        except Exception as exc:
            raise CSVValidationError(
                'Файл повреждён или не является настоящим .xls. '
                'Пересохраните его в Excel как «Книга Excel 97-2003 (.xls)».'
            ) from exc
        sheet = wb.sheet_by_index(0)
        self._validate_sheet_dimensions(sheet.nrows, sheet.ncols)

        def iter_rows() -> Iterator[tuple]:
            for row_idx in range(sheet.nrows):
                row = []
                for col_idx in range(sheet.ncols):
                    value = sheet.cell_value(row_idx, col_idx)
                    # xlrd отдаёт все числа как float; целые приводим к int
                    # (12345.0 → 12345), чтобы артикулы/количества совпадали
                    # с поведением .xlsx и CSV.
                    if isinstance(value, float) and value.is_integer():
                        value = int(value)
                    row.append(value)
                yield tuple(row)

        try:
            return self._rows_to_dicts(iter_rows())
        finally:
            release_resources = getattr(wb, 'release_resources', None)
            if release_resources is not None:
                release_resources()

    @staticmethod
    def _validate_sheet_dimensions(row_count: int | None, column_count: int | None) -> None:
        """Reject sparse sheets whose declared range would allocate giant rows."""
        max_rows = datasource_limit('DATASOURCE_IMPORT_MAX_ROWS')
        max_columns = datasource_limit('DATASOURCE_IMPORT_MAX_COLUMNS')
        max_cells = datasource_limit('DATASOURCE_IMPORT_MAX_CELLS')
        rows = max(0, int(row_count or 0))
        columns = max(0, int(column_count or 0))
        if rows > max_rows:
            raise CSVValidationError(
                'Количество строк превышает допустимый лимит '
                f'{max_rows}.'
            )
        if columns > max_columns:
            raise CSVValidationError(
                'Количество колонок превышает допустимый лимит '
                f'{max_columns}.'
            )
        if rows and columns > max_cells // rows:
            raise CSVValidationError(
                'Количество ячеек в заявленном диапазоне листа превышает лимит '
                f'{max_cells} ячеек.'
            )

    def _iter_limited_rows(
        self,
        rows: Iterable[Sequence],
    ) -> Iterator[tuple]:
        max_rows = datasource_limit('DATASOURCE_IMPORT_MAX_ROWS')
        max_columns = datasource_limit('DATASOURCE_IMPORT_MAX_COLUMNS')
        max_cells = datasource_limit('DATASOURCE_IMPORT_MAX_CELLS')
        total_cells = 0
        for row_number, source_row in enumerate(rows, start=1):
            if row_number > max_rows:
                raise CSVValidationError(
                    'Количество строк превышает допустимый лимит '
                    f'{max_rows}.'
                )
            row = tuple(source_row)
            if len(row) > max_columns:
                raise CSVValidationError(
                    'Количество колонок превышает допустимый лимит '
                    f'{max_columns}.',
                )
            total_cells += len(row)
            if total_cells > max_cells:
                raise CSVValidationError(
                    'Количество ячеек превышает допустимый лимит '
                    f'{max_cells}.'
                )
            yield row

    def _rows_to_dicts(self, rows: Iterable[Sequence]) -> list[dict]:
        row_iterator = iter(self._iter_limited_rows(rows))
        buffered_rows = list(islice(
            row_iterator,
            self.MAX_HEADER_SCAN_ROWS + self.MAX_HEADER_DEPTH + self.HEADER_DATA_SAMPLE_ROWS,
        ))
        if not buffered_rows:
            return []
        header_start, header_depth, headers = self._detect_header(buffered_rows)
        result = []
        last_group_values = {}
        data_rows = chain(buffered_rows[header_start + header_depth:], row_iterator)
        for row in data_rows:
            row_dict = {}
            has_any_value = False
            for i, header in enumerate(headers):
                if not header:
                    continue
                value = row[i] if i < len(row) else ''
                value = str(value).strip() if value is not None else ''
                if value:
                    has_any_value = True
                row_dict[header] = value
            if not has_any_value:
                continue

            for group_key in ('brand', 'category'):
                value = row_dict.get(group_key, '')
                if value:
                    last_group_values[group_key] = value
                elif row_dict.get('article') or row_dict.get('name'):
                    row_dict[group_key] = last_group_values.get(group_key, '')
            result.append(row_dict)
        return result

    def _detect_header(self, rows: list[tuple]) -> tuple[int, int, list[str]]:
        best = None
        scan_limit = min(len(rows), self.MAX_HEADER_SCAN_ROWS)
        for start in range(scan_limit):
            for depth in range(1, self.MAX_HEADER_DEPTH + 1):
                if start + depth > len(rows):
                    continue
                headers = self._build_headers(rows[start:start + depth])
                found_required = {header for header in headers if header in self.REQUIRED_COLUMNS}
                if not found_required:
                    continue
                data_score = self._score_data_rows(rows[start + depth:start + depth + 5], headers)
                score = len(found_required) * 100 + data_score - start * 2 - (depth - 1) * 5
                candidate = (score, start, depth, headers)
                if best is None or candidate[0] > best[0]:
                    best = candidate

        if best is None:
            headers = [str(h).strip() if h is not None else '' for h in rows[0]]
            return 0, 1, headers
        _, start, depth, headers = best
        return start, depth, headers

    def _build_headers(self, header_rows: list[tuple]) -> list[str]:
        width = max((len(row) for row in header_rows), default=0)
        headers = []
        parent_labels = [''] * width
        for col_idx in range(width):
            labels = []
            for row in header_rows:
                label = str(row[col_idx]).strip() if col_idx < len(row) and row[col_idx] is not None else ''
                if label:
                    labels.append(label)
                    parent_labels[col_idx] = label
            if not labels and parent_labels[col_idx]:
                labels.append(parent_labels[col_idx])
            headers.append(self._canonical_column(' '.join(labels)))
        return headers

    def _canonical_column(self, value: str) -> str:
        normalized = ' '.join(str(value).lower().replace('\n', ' ').split())
        normalized = normalized.replace('ё', 'е')
        if not normalized:
            return ''
        exact = self.COLUMN_ALIASES.get(normalized)
        if exact:
            return exact

        checks = [
            ('article', ('артикул', 'номер производ', 'номер детали', 'part number', 'sku')),
            ('brand', ('производитель', 'бренд', 'марка', 'brand')),
            ('name', ('номенклатура', 'наименование', 'название', 'товар', 'описание товара', 'name')),
            ('price', ('средняя цена', 'цена', 'стоимость', 'price')),
            (
                'stock_qty',
                (
                    'конечный остаток количество',
                    'остаток количество',
                    'остаток',
                    'количество',
                    'кол-во',
                    'qty',
                    'stock',
                ),
            ),
            ('category', ('категория', 'группа', 'раздел', 'category')),
            ('condition', ('состояние', 'condition')),
            ('oem_numbers', ('oem',)),
            ('cross_numbers', ('кросс', 'cross')),
            ('description', ('описание', 'description')),
        ]
        for key, markers in checks:
            if any(marker in normalized for marker in markers):
                return key
        return normalized

    def _score_data_rows(self, rows: list[tuple], headers: list[str]) -> int:
        score = 0
        for row in rows:
            row_map = {
                headers[i]: row[i]
                for i in range(min(len(headers), len(row)))
                if headers[i] in self.REQUIRED_COLUMNS and row[i] not in (None, '')
            }
            score += len(row_map)
            if row_map.get('article') and row_map.get('name'):
                score += 2
            if self._looks_numeric(row_map.get('price')):
                score += 1
            if self._looks_numeric(row_map.get('stock_qty')):
                score += 1
        return score

    def _looks_numeric(self, value) -> bool:
        if value in (None, ''):
            return False
        try:
            Decimal(str(value).replace(',', '.').replace(' ', ''))
        except InvalidOperation:
            return False
        return True

    COLUMN_ALIASES = {
        'артикул': 'article',
        'номер производ.': 'article',
        'номер производителя': 'article',
        'название': 'name',
        'наименование': 'name',
        'номенклатура': 'name',
        'товар': 'name',
        'цена': 'price',
        'стоимость': 'price',
        'остаток': 'stock_qty',
        'количество': 'stock_qty',
        'кол-во': 'stock_qty',
        'бренд': 'brand',
        'производитель': 'brand',
        'марка': 'brand',
        'категория': 'category',
        'состояние': 'condition',
        'oem': 'oem_numbers',
        'кроссы': 'cross_numbers',
        'кросс-номера': 'cross_numbers',
        'описание': 'description',
    }

    def _normalize(self, rows: list[dict]) -> list[dict]:
        if not rows:
            return []

        # Переименовываем колонки по алиасам для каждой строки
        normalized_rows = []
        for row in rows:
            mapped_row = {}
            for k, v in row.items():
                if k is None:
                    continue
                mapped_key = self._canonical_column(k)
                mapped_row[mapped_key] = v
            normalized_rows.append(mapped_row)

        mapped_headers = set(normalized_rows[0].keys()) if normalized_rows else set()
        missing = [col for col in self.REQUIRED_COLUMNS if col not in mapped_headers]
        if missing:
            raise CSVValidationError(f'Отсутствуют обязательные колонки: {", ".join(missing)}')

        result = []
        for i, row in enumerate(normalized_rows, start=2):
            normalized = {k: (v.strip() if isinstance(v, str) else v) for k, v in row.items()}
            if not any(normalized.get(col) for col in self.REQUIRED_COLUMNS):
                continue
            if not normalized.get('article') or not normalized.get('name'):
                continue

            # Если цена пустая, пропускаем или ставим 0? Поставим 0 или выкинем ошибку
            price_val = str(normalized.get('price') or 0).replace(',', '.').replace(' ', '')
            try:
                price = Decimal(price_val)
            except InvalidOperation:
                raise CSVValidationError(f'Строка {i}: некорректная цена "{normalized["price"]}"')
            try:
                stock_val = str(normalized.get('stock_qty') or 0).replace(',', '.').replace(' ', '')
                stock_qty = int(float(stock_val))
            except (ValueError, TypeError):
                raise CSVValidationError(f'Строка {i}: некорректное количество "{normalized["stock_qty"]}"')

            result.append({
                'uuid': None,
                'article': normalized['article'],
                'name': normalized['name'],
                'brand': normalized.get('brand', ''),
                'price': str(price),
                'stock_qty': stock_qty,
                'category': normalized.get('category', ''),
                'condition': normalized.get('condition', 'new'),
                'oem_numbers': normalized.get('oem_numbers', ''),
                'cross_numbers': normalized.get('cross_numbers', ''),
                'description': normalized.get('description', ''),
            })
        return result
